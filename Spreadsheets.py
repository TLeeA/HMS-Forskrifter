#! /usr/bin/python3

# Use Python and openpyxl to create Excel workbook with several sheets, and
# hopefully ... import the csv files into individual sheets

# help found at:
# https://linuxconfig.org/how-to-manipulate-excel-spreadsheets-with-python-and-openpyxl

# testing first:

from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
#spreadsheet = wb.active

# Populate the cell
#spreadsheet.cell(row=1, column=1, value='Hello World')

wb.remove(wb['Sheet'])
new_sheet = wb.create_sheet('Forskrifter->',1)
spreadsheet = wb.active
# Populate the cell
spreadsheet.cell(row=1, column=1, value='HMS forskrifter for Petroleumsnæringen i Norge')
spreadsheet.cell(row=3, column=1, value='Ingen garanti for dette produkt!')
spreadsheet.cell(row=4, column=1, value='Terje L. Andersen, April 2022')

new_sheet = wb.create_sheet('ramme',2)
#spreadsheet = wb
#spreadsheet.cell(row=1, column=1, value='Ramme her')
new_sheet = wb.create_sheet('styrings',3)
new_sheet = wb.create_sheet('innretnings',4)
new_sheet = wb.create_sheet('aktivitets',5)

shts = wb.sheetnames
print(shts)

wb.save('HMS-Forskrifter.xlsx')


# this seems to only work for one single sheet in the workbook...
# https://stackoverflow.com/questions/68290557/how-can-i-write-a-csv-file-to-an-excel-sheet-using-openpyxl
#import pandas as pd
#df = pd.read_csv('rammeforskriften.csv','\\')
#df.to_excel('HMS-Forskrifter.xlsx', sheet_name='ramme', header=1, index=False)


# Try this in stead:
# https://www.tutorialguruji.com/python/python-csv-to-excel-sheet-by-sheet-with-openpyxl-not-optimized/

#import openpyxl
import csv
import os

#wb = openpyxl.load_workbook('HMS-Forskrifter.xlsx')
wb = load_workbook('HMS-Forskrifter.xlsx')

sheets = [
    ('rammeforskriften.csv', 'ramme'),
    ('styringsforskriften.csv', 'styrings'),
    ('innretningsforskriften.csv', 'innretnings'),
    ('aktivitetsforskriften.csv', 'aktivitets')
]

for filename, sheet in sheets:
    with open(filename, newline='') as f_input:
        ws = wb[sheet]
        
        for rowy, row in enumerate(csv.reader(f_input, delimiter='\\'), start=1):    # make start first row number
            for colx, value in enumerate(row, start=1):                             # make start first col number    
                ws.cell(column=colx, row=rowy, value=value)
        
wb.save('HMS-Forskrifter.xlsx')

